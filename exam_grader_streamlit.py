"""
Exam Grader Pro - Streamlit Web App
A professional tool for grading multiple-choice exams with intelligent answer parsing.

Author: Jorge B. Cevallos
"""

import streamlit as st
import re
import io
import pandas as pd
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ============================================================================
# CONFIGURATION
# ============================================================================

st.set_page_config(
    page_title="Exam Grader Pro",
    page_icon="✓",
    layout="wide"  # Use full screen width
)

# ============================================================================
# PASSWORD GATE
# ============================================================================

def check_password():
    if "authenticated" not in st.session_state:
        st.session_state.authenticated = False

    if st.session_state.authenticated:
        return True

    st.markdown("## ✓ Exam Grader Pro")
    st.markdown("---")
    pwd = st.text_input("Password", type="password", placeholder="Enter password")
    if st.button("Login", type="primary"):
        if pwd.lower() == "portafolio":
            st.session_state.authenticated = True
            st.rerun()
        else:
            st.error("Incorrect password.")
    return False

if not check_password():
    st.stop()

# ============================================================================
# HELPER FUNCTIONS
# ============================================================================

def extract_letters(input_text):
    """Extract only letters from input text, removing numbers, punctuation, and whitespace."""
    if not input_text:
        return ""
    # Remove all non-letter characters and convert to uppercase
    return re.sub(r'[^A-Z]', '', input_text.upper())


def grade_exam(answer_key, student_answer, grading_scale):
    """
    Compare student answer against answer key and calculate grade.
    
    Returns:
        tuple: (grade, correct_count, incorrect_count, total_questions, percentage, comparison_data)
    """
    # Extract letters only
    key_letters = extract_letters(answer_key)
    student_letters = extract_letters(student_answer)
    
    if not key_letters or not student_letters:
        return None, None, None, None, None, None
    
    # Get the shorter length for comparison
    total_questions = min(len(key_letters), len(student_letters))
    
    # Compare answers
    correct_count = 0
    comparison_data = []
    
    for i in range(total_questions):
        is_correct = key_letters[i] == student_letters[i]
        if is_correct:
            correct_count += 1
        
        comparison_data.append({
            'question': i + 1,
            'key': key_letters[i],
            'student': student_letters[i],
            'correct': is_correct
        })
    
    # Calculate metrics
    incorrect_count = total_questions - correct_count
    percentage = (correct_count / total_questions * 100) if total_questions > 0 else 0
    grade = (correct_count / total_questions * grading_scale) if total_questions > 0 else 0
    
    return grade, correct_count, incorrect_count, total_questions, percentage, comparison_data


# ============================================================================
# BATCH MODE HELPERS
# ============================================================================

def parse_batch_entry(raw_text):
    """Parse a 2-line Moodle entry: line 1 = name (markdown or plain), line 2 = answers."""
    lines = [l.strip() for l in raw_text.strip().splitlines() if l.strip()]
    if len(lines) < 1:
        return None, None
    name_line  = lines[0]
    answer_line = lines[1] if len(lines) >= 2 else ""
    md_match = re.match(r'\[([^\]]+)\]', name_line)
    name = md_match.group(1).strip().upper() if md_match else name_line.strip().upper()
    answers = extract_letters(answer_line) if answer_line else ""
    return name, answers


def format_nota(value):
    """Integer if whole, else 2-decimal with comma (e.g. 8 → '8', 9.67 → '9,67')."""
    if value == int(value):
        return str(int(value))
    return f"{value:.2f}".replace(".", ",")


def grade_batch_student(key_letters, student_letters, grading_scale):
    """Grade one student against the key. Returns result dict."""
    n = len(key_letters)
    obs_flag = None
    if len(student_letters) > n:
        obs_flag = f"{len(student_letters)} caracteres enviados; se calificaron los primeros {n}"
        student_letters = student_letters[:n]
    correct   = sum(1 for i in range(n)
                    if i < len(student_letters) and student_letters[i] == key_letters[i])
    wrong_pos = [i + 1 for i in range(n)
                 if not (i < len(student_letters) and student_letters[i] == key_letters[i])]
    grade = (correct / n * grading_scale) if n > 0 else 0
    return {
        "correct":        correct,
        "total":          n,
        "grade":          grade,
        "nota":           format_nota(grade),
        "wrong_positions": wrong_pos,
        "obs_flag":       obs_flag,
    }


def build_batch_excel(entries, key_length, grading_scale):
    """Build Excel workbook from batch entries list. Returns bytes."""
    sorted_entries = sorted(entries, key=lambda x: x["name"])

    # Detect duplicate answer strings
    answer_map = {}
    for e in sorted_entries:
        answer_map.setdefault(e["answers"], []).append(e["name"])

    wb = Workbook()
    ws = wb.active
    ws.title = "Calificaciones"

    hdr_font  = Font(name="Arial", bold=True, color="FFFFFF", size=11)
    hdr_fill  = PatternFill("solid", start_color="2F4F7F")
    hdr_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    yellow    = PatternFill("solid", start_color="FFFF00")
    red_f     = PatternFill("solid", start_color="FF6B6B")
    alt       = PatternFill("solid", start_color="F2F2F2")
    no_fill   = PatternFill(fill_type=None)
    center_a  = Alignment(horizontal="center", vertical="center")
    left_a    = Alignment(horizontal="left",   vertical="center", wrap_text=True)
    thin      = Side(style="thin", color="CCCCCC")
    bdr       = Border(left=thin, right=thin, top=thin, bottom=thin)

    headers    = ["APELLIDOS Y NOMBRES", f"ACIERTOS/{key_length}", "NOTA", "OBSERVACIÓN"]
    col_widths = [42, 14, 10, 65]

    for col, (h, w) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font      = hdr_font
        cell.fill      = hdr_fill
        cell.alignment = hdr_align
        cell.border    = bdr
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.row_dimensions[1].height = 30

    for i, e in enumerate(sorted_entries, 2):
        r = e["result"]
        obs_parts = []
        group = answer_map.get(e["answers"], [])
        if len(group) > 1:
            others = [n for n in group if n != e["name"]]
            obs_parts.append("⚠ Idéntica: " + " / ".join(others))
        if r["obs_flag"]:
            obs_parts.append(r["obs_flag"])
        obs     = " | ".join(obs_parts)
        is_dup  = len(group) > 1
        is_zero = r["grade"] == 0

        row_data = [e["name"], r["correct"], r["nota"], obs]
        for col, val in enumerate(row_data, 1):
            cell = ws.cell(row=i, column=col, value=val)
            cell.font      = Font(name="Arial", size=10)
            cell.border    = bdr
            cell.alignment = center_a if col in (2, 3) else left_a
            cell.fill      = red_f if is_zero else (yellow if is_dup else (alt if i % 2 == 0 else no_fill))
        ws.row_dimensions[i].height = 18

    ws.freeze_panes = "A2"
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


# ============================================================================
# STREAMLIT APP
# ============================================================================

def main():
    # Header
    col_title1, col_title2 = st.columns([3, 1])
    with col_title1:
        st.title("✓ Exam Grader Pro")
        st.caption("Generate accurate grades from multiple-choice exams with intelligent answer parsing")
    with col_title2:
        with st.expander("📖 How to use"):
            st.markdown("""
            **Single:** key + one student → instant grade.

            **Batch:** set key, paste Moodle entries one by one, download Excel when done.
            Each entry = name line + answers line (copy directly from Moodle).
            """)

    st.markdown("---")
    tab1, tab2 = st.tabs(["📝 Single Grader", "📦 Batch Mode"])

    # ── TAB 1: existing single-grader ────────────────────────────────────────
    with tab1:
        left_col, right_col = st.columns([1, 1])

        with left_col:
            # Configuration (compact)
            scale_option = st.selectbox(
                "⚙️ Grading Scale",
                ["10-point scale", "100-point scale", "20-point scale", "5-point scale", "Custom"]
            )
            
            if scale_option == "Custom":
                grading_scale = st.number_input("Custom Scale", min_value=1, max_value=1000, value=10, step=1)
            else:
                grading_scale = int(scale_option.split('-')[0])
            
            # Answer Key (compact)
            answer_key = st.text_area(
                "🔑 Answer Key",
                height=60,
                placeholder="ADCABCBADCBA...",
                key="answer_key"
            )
            
            # Student Answer (compact)
            student_answer = st.text_area(
                "📝 Student's Answer",
                height=120,
                placeholder="Paste answer in any format",
                key="student_answer"
            )
            
            # Buttons (compact)
            col1, col2 = st.columns(2)
            with col1:
                calculate_button = st.button("🚀 Calculate", type="primary", use_container_width=True)
            with col2:
                if st.button("Clear", use_container_width=True):
                    st.rerun()

        # Right column - Results (always starts at top)
        with right_col:
            # Add minimal spacing to align with grading scale label
            st.markdown("<div style='height: 28px;'></div>", unsafe_allow_html=True)
            
            # Process grading
            if calculate_button:
                if not answer_key or not student_answer:
                    st.error("❌ Please enter both answer key and student answer.")
                else:
                    # Extract letters for validation
                    key_letters = extract_letters(answer_key)
                    student_letters = extract_letters(student_answer)
                    
                    if not key_letters:
                        st.error("❌ Could not extract valid letters from answer key.")
                    elif not student_letters:
                        st.error("❌ Could not extract valid letters from student answer.")
                    else:
                        # Check for length mismatch
                        if len(key_letters) != len(student_letters):
                            st.warning(f"⚠️ Length mismatch: Key={len(key_letters)}, Student={len(student_letters)}")
                        
                        # Calculate grade
                        grade, correct, incorrect, total, percentage, comparison = grade_exam(
                            answer_key, student_answer, grading_scale
                        )
                        
                        if grade is not None:
                            # GRADE - BIG AND PROMINENT AT TOP
                            st.markdown(f"""
                            <div style="background: linear-gradient(135deg, #4299e1 0%, #667eea 100%); 
                                        padding: 32px; 
                                        border-radius: 12px; 
                                        text-align: center; 
                                        color: white;
                                        margin-bottom: 20px;
                                        box-shadow: 0 4px 12px rgba(66, 153, 225, 0.3);">
                                <div style="font-size: 11px; font-weight: 600; text-transform: uppercase; letter-spacing: 1.5px; opacity: 0.85; margin-bottom: 12px;">
                                    FINAL GRADE
                                </div>
                                <div style="font-size: 64px; font-weight: 700; letter-spacing: -2px; line-height: 1;">
                                    {grade:.2f} / {grading_scale}
                                </div>
                            </div>
                            """, unsafe_allow_html=True)
                            
                            # Compact stats - small and uniform
                            st.markdown(f"""
                            <div style="display: flex; justify-content: space-around; margin-bottom: 16px; padding: 12px; background: #f9fafb; border-radius: 8px;">
                                <div style="text-align: center;">
                                    <div style="font-size: 11px; color: #6b7280; font-weight: 600; margin-bottom: 4px;">✓ CORRECT</div>
                                    <div style="font-size: 18px; font-weight: 700; color: #10b981;">{correct}</div>
                                </div>
                                <div style="text-align: center;">
                                    <div style="font-size: 11px; color: #6b7280; font-weight: 600; margin-bottom: 4px;">✗ INCORRECT</div>
                                    <div style="font-size: 18px; font-weight: 700; color: #ef4444;">{incorrect}</div>
                                </div>
                                <div style="text-align: center;">
                                    <div style="font-size: 11px; color: #6b7280; font-weight: 600; margin-bottom: 4px;">TOTAL</div>
                                    <div style="font-size: 18px; font-weight: 700; color: #374151;">{total}</div>
                                </div>
                                <div style="text-align: center;">
                                    <div style="font-size: 11px; color: #6b7280; font-weight: 600; margin-bottom: 4px;">PERCENTAGE</div>
                                    <div style="font-size: 18px; font-weight: 700; color: #374151;">{percentage:.1f}%</div>
                                </div>
                            </div>
                            """, unsafe_allow_html=True)
                            
                            # Copy grade - compact and centered
                            col_left_copy, col_copy, col_right_copy = st.columns([1, 1, 1])
                            with col_copy:
                                st.code(f"{grade:.2f}", language=None)
                                st.markdown("<p style='text-align: center; font-size: 11px; color: #6b7280; margin-top: -8px;'>👆 Copy to LMS</p>", unsafe_allow_html=True)
                            
                            st.markdown("<div style='margin: 24px 0;'></div>", unsafe_allow_html=True)
                            
                            # Detailed comparison - centered title and centered table (NO LINE ABOVE)
                            st.markdown("<h3 style='text-align: center; margin-bottom: 16px;'>Detailed Comparison</h3>", unsafe_allow_html=True)
                            
                            # Create DataFrame
                            comparison_df = pd.DataFrame(comparison)
                            comparison_df['Q'] = comparison_df['question']
                            comparison_df['Key'] = comparison_df['key']
                            comparison_df['Student'] = comparison_df['student']
                            comparison_df['✓/✗'] = comparison_df['correct'].apply(lambda x: '✓' if x else '✗')
                            
                            # Center the table using columns with padding
                            col_left, col_table, col_right = st.columns([0.5, 2, 0.5])
                            
                            with col_table:
                                # Display table centered
                                st.dataframe(
                                    comparison_df[['Q', 'Key', 'Student', '✓/✗']],
                                    use_container_width=True,
                                    hide_index=True,
                                    height=300
                                )
            else:
                # Placeholder
                st.info("👈 Enter data and click Calculate to see grade here")

    # ── TAB 2: batch mode ────────────────────────────────────────────────────
    with tab2:
        st.markdown("### 📦 Batch Mode")
        st.caption("Copy each student's Moodle block (name line + answers line) and add them one by one.")

        # Config row
        bc1, bc2, bc3 = st.columns([2, 1, 1])
        with bc1:
            batch_key = st.text_input(
                "🔑 Answer Key",
                key="batch_key",
                placeholder="BCBABBBABCBCAAABCBCCACCCABACAB"
            )
        with bc2:
            b_scale_opt = st.selectbox(
                "⚙️ Scale",
                ["10-point", "100-point", "20-point", "5-point", "Custom"],
                key="b_scale_opt"
            )
            if b_scale_opt == "Custom":
                b_scale = st.number_input("Custom", min_value=1, max_value=1000, value=10, key="b_scale_custom")
            else:
                b_scale = int(b_scale_opt.split("-")[0])
        with bc3:
            st.markdown("<div style='height:28px'></div>", unsafe_allow_html=True)
            st.metric("Entries added", len(st.session_state.get("batch_entries", [])))

        st.markdown("---")
        col_input, col_list = st.columns([1, 1])

        # ── Left: paste & add ────────────────────────────────────────────────
        with col_input:
            st.markdown("**Paste Moodle entry (2 lines):**")
            entry_text = st.text_area(
                "entry",
                height=130,
                key="batch_entry_text",
                label_visibility="collapsed",
                placeholder="[APELLIDOS NOMBRES](https://eidiomas.espam...)\nBCBABBBABCBCAAABCBCCACCCABACAB"
            )
            btn_add, btn_clear = st.columns(2)
            with btn_add:
                add_clicked = st.button("➕ Add Entry", type="primary", use_container_width=True)
            with btn_clear:
                if st.button("🗑 Clear All", use_container_width=True):
                    st.session_state.batch_entries = []
                    st.rerun()

            if add_clicked:
                if not batch_key:
                    st.error("Set an answer key first.")
                elif not entry_text.strip():
                    st.error("Paste an entry first.")
                else:
                    name, answers = parse_batch_entry(entry_text)
                    if not name:
                        st.error("Could not parse name from entry.")
                    else:
                        key_letters = extract_letters(batch_key)
                        result      = grade_batch_student(key_letters, answers, b_scale)
                        if "batch_entries" not in st.session_state:
                            st.session_state.batch_entries = []
                        existing_names = [e["name"] for e in st.session_state.batch_entries]
                        if name in existing_names:
                            st.warning(f"⚠ '{name}' is already in the batch.")
                        else:
                            st.session_state.batch_entries.append({
                                "name":    name,
                                "answers": answers,
                                "result":  result,
                            })
                            st.success(f"✓ {name}  →  {result['nota']}")
                            st.rerun()

        # ── Right: running list + download ───────────────────────────────────
        with col_list:
            entries = st.session_state.get("batch_entries", [])
            if entries:
                # Detect duplicates for preview table
                answer_map_preview = {}
                for e in entries:
                    answer_map_preview.setdefault(e["answers"], []).append(e["name"])

                rows = []
                for e in entries:
                    dup = len(answer_map_preview.get(e["answers"], [])) > 1
                    rows.append({
                        "NAME":    e["name"],
                        "CORRECT": e["result"]["correct"],
                        "NOTA":    e["result"]["nota"],
                        "⚠":      "⚠" if dup else "",
                    })

                st.markdown(f"**Batch preview ({len(entries)} entries):**")
                st.dataframe(
                    pd.DataFrame(rows),
                    use_container_width=True,
                    hide_index=True,
                    height=220
                )

                if batch_key:
                    key_letters  = extract_letters(batch_key)
                    excel_bytes  = build_batch_excel(entries, len(key_letters), b_scale)
                    filename     = f"Calificaciones_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
                    st.download_button(
                        label="📥 Download Excel",
                        data=excel_bytes,
                        file_name=filename,
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        use_container_width=True,
                        type="primary"
                    )
                else:
                    st.warning("Set the answer key above to enable Excel export.")
            else:
                st.info("No entries yet. Paste and add entries on the left.")

    # Footer
    st.markdown("---")
    st.markdown("""
    <div style="text-align: center; color: #6b7280; font-size: 13px; padding: 16px 0;">
        <div style="margin-bottom: 8px;"><strong>Smart Format Detection</strong></div>
        <div>Created by <strong style="color: #374151;">Jorge B. Cevallos</strong></div>
    </div>
    """, unsafe_allow_html=True)


if __name__ == "__main__":
    main()
